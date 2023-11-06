import json
import os

import openpyxl.utils
import openpyxl
import pandas as pd
from itertools import product
import ast
#{'Feuille': 'Réél Biais stratégie', 'Colonne': 'Date début', 'Opérateur': '=', 'Valeur': 'ere'}
from PySide2.QtWidgets import QLineEdit, QLabel, QTextEdit
from openpyxl.worksheet.hyperlink import Hyperlink

def extraire_label_dico_scroll_area_special(scroll_area):
    liste_dico=[]
    liste_widget=[]

    for i in range(scroll_area.count()):
        widget = scroll_area.itemAt(i).widget()
        if widget:
            # Vérifier le type du widget (par exemple, QLineEdit, QLabel, QTextEdit, etc.)
            if isinstance(widget, QLineEdit):
                texte_critère = widget.text()
            elif isinstance(widget, QLabel):
                texte_critère = widget.text()
            elif isinstance(widget, QTextEdit):
                texte_critère = widget.toPlainText()
            liste_dico.append(ast.literal_eval(texte_critère))
            print(texte_critère)
            liste_widget.append(widget)

    return {'liste_dico':liste_dico,'liste_widget':liste_widget}

def _traitement_critère(chemin,dico_critère):
    sheet_name=dico_critère["Feuille"]
    colonne=dico_critère['Colonne']
    opérateur=dico_critère["Opérateur"]
    value=dico_critère["Valeur"]
    df=pd.read_excel(chemin,sheet_name=sheet_name)
    if opérateur=="=":
        filtered_df = df[df[colonne] == value]
    elif opérateur==">":
        filtered_df = df[df[colonne] > value]
    elif opérateur=="<":
        filtered_df = df[df[colonne] < value]
    elif opérateur=="<=":
        filtered_df = df[df[colonne] <= value]
    elif opérateur==">=":
        filtered_df = df[df[colonne] >= value]
    elif opérateur=="ne contient pas":
        filtered_df = df[~df[colonne].str.contains(value)]
    elif opérateur=="contient":
        filtered_df = df[df[colonne].str.contains(value)]

    serie_number_trade=filtered_df["Trade"]
    return serie_number_trade




def traitement_critères(liste_dico_critère,chemin):
    #Trie des données
    liste_série_nb_trades=[]
    for dico in liste_dico_critère:
        série=_traitement_critère(chemin,dico)
        liste_série_nb_trades.append(série)
    #fusion des séries Trade pour filtrer datas

    serie_fusionnee = liste_série_nb_trades[0]
    if len(liste_série_nb_trades) > 1:
        for série in liste_série_nb_trades[1:]:
            serie_fusionnee = pd.merge(serie_fusionnee, série, on='Trade')
    return serie_fusionnee


#fonction pour extraire chaque série individuellement du df
def _extr_data(chemin,data,serie_nb_trade):
    with open("C:\\Users\\Baptiste\\Documents\\GUI_ext_tri_datas\\données.json", 'r') as file:
        json_data = json.load(file)

    feuille=data["Feuille"]
    colonne=data["Colonne"]
    datas_screenshot=json_data["valeurs_defaults"]["Screen_words"]
    #---> Faire condition ici pour virer les colonne screenshots
    for word in datas_screenshot:
        if word in colonne:
            return(False,[feuille,colonne])
    df=pd.read_excel(chemin,sheet_name=feuille)
    serie=pd.merge(serie_nb_trade,df,on="Trade")
    serie_filtred=serie[colonne]
    return (True,serie_filtred)

def extraire_datas(chemin,liste_datas_extr,serie_nb_trade):
    listes_séries_extr=[serie_nb_trade]
    colonnes_screenshot=[]
    for data in liste_datas_extr:
        serie=_extr_data(chemin,data,serie_nb_trade)
        if serie[0]:
            listes_séries_extr.append(serie[1])
        else:
            colonnes_screenshot.append(serie[1])
    df_concat=pd.concat(listes_séries_extr,axis=1)
    return (df_concat,colonnes_screenshot)

#Permet de changer une colonne en hyperlien si il contient mot clé liste=[[colonne,feuille]]
def écrire_colonne_hyperlink(chemin_fichier_racine,liste_screenshots,chemin_final,serie_nber_trade,name_sheet):
    #json data
    with open("C:\\Users\\Baptiste\\Documents\\GUI_ext_tri_datas\\données.json","r") as json_file:
        data=json.load(json_file)
    liste_words_trade=data["valeurs_defaults"]["col_ref_words"]
    #serie_nber_trade est la serie de nuber trade du df
    workbook_dst=openpyxl.load_workbook(chemin_final)
    workbook_read=openpyxl.load_workbook(chemin_fichier_racine)
    ws_ecrire=workbook_dst[name_sheet]

    #Ca c'est bon
    for entete in liste_screenshots:
        colonne = openpyxl.utils.get_column_letter(ws_ecrire.max_column +1)
        ws_ecrire[f"{colonne}1"] = entete[1]  # Écrivez l'entête dans la première ligne

   #Boucle pour récupere l'emplacement de la colonne de screenshot +la colonne trade+la feuille
    #{Nom col:{emplacement coll,worksheet,emplacement_trade}}
    liste_emplacement_screenshot={}
    #OK FONCTION VERIFIE
    for emplacement_screen in liste_screenshots:
        ws=workbook_read[emplacement_screen[0]]
        emplacement_trade_trouvé = False
        emplacement_screen_trouvé=False
        emplacement_trade=""
        emplacement_screenshot=""
        for col_idx, col_name_cell in enumerate(ws[1], start=1):

            col_name = col_name_cell.value
            for word in liste_words_trade:
                if word in col_name:
                    emplacement_trade=openpyxl.utils.get_column_letter(col_idx)
                    emplacement_trade_trouvé=True
            if col_name ==emplacement_screen[1]:
                emplacement_colonne = openpyxl.utils.get_column_letter(col_idx)
                emplacement_screenshot=True
            if emplacement_screenshot and emplacement_trade_trouvé:
                liste_emplacement_screenshot[col_name]={"emplacement_colonne":emplacement_colonne,
                                                    "ws":ws,"emplacement_trade":emplacement_trade}
                break

        # {Nom col:{emplacement coll,worksheet,emplacement_trade}}
        for nom_colonne_screnshot,dico_emplacement in liste_emplacement_screenshot.items():
            ws=dico_emplacement['ws']
            lettre_coll_trade=dico_emplacement['emplacement_trade']
            emplacement_colonne=dico_emplacement["emplacement_colonne"]
            # Parcourez la colonne ligne par ligne
            for cellule in ws[lettre_coll_trade]:
                valeur = cellule.value
                #le numéro de trade match avec la serie contenant les trades
                if valeur in serie_nber_trade.values:
                    #récupérer la valeur et l'hyperlien du screenshot
                    try:
                        cell_data_hyperlink=ws[emplacement_colonne][cellule.row-1]
                    except:pass
                    print(emplacement_colonne,cellule.row)
                    if cell_data_hyperlink.hyperlink:
                        hyperlink=cell_data_hyperlink.hyperlink.target
                    else:hyperlink=""
                    value=cell_data_hyperlink.value

                    #appliquer à la celulle
                    for cell in ws_ecrire["A"]:
                        if cell.value==valeur:
                            row_hyperlink_write=cell.row

                    for e in ws_ecrire[1]:


                        if e.value == nom_colonne_screnshot:
                            nom_de_colonne = openpyxl.utils.get_column_letter(e.column)

                            break

                    cell_hyperlien=ws_ecrire[f"{nom_de_colonne}{row_hyperlink_write}"]
                    repertoire = os.path.dirname(chemin_fichier_racine)

                    if "https:" not in hyperlink and "C:/" not in hyperlink:

                    #régler le problleme d'hyperlien
                        cell_hyperlien.hyperlink=repertoire+"/"+hyperlink
                    else:
                        cell_hyperlien.hyperlink=hyperlink
                    cell_hyperlien.value=value
    workbook_dst.save(chemin_final)








# def change_colonne_to_hyperlink(chemin):
    # Telecharger datas:
    # Emplacement du répertoire du script Python
    #emplacement_script = os.path.dirname(os.path.abspath(__file__))
    # Nom de votre fichier JSON
    #nom_fichier_json = "données.json"
    # Construction du chemin relatif vers le répertoire parent
    #chemin_json = os.path.join(emplacement_script, nom_fichier_json)
    #with open(chemin_json, "r") as fichier_json:
    #    data = json.load(fichier_json)
    #screen_words = data["valeurs_defaults"]["Screen_words"]
    #print(screen_words)
    # Ouvrez le classeur Excel existant
    #wb = openpyxl.load_workbook(chemin)

    # Accédez à la feuille de calcul (remplacez 'Feuille1' par le nom de votre feuille)
    #ws = wb['Sheet1']

    # Parcourez les colonnes pour trouver celles qui contiennent "Screen" dans leur nom
    #for col in ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=1, max_row=1):
    #    for cell in col:
    #        for word in screen_words:
    #           if word in cell.value:
    #               # Parcourez les lignes de cette colonne et ajoutez un hyperlien à chaque cellule
    #               for row_cell in ws.iter_rows(min_row=2, min_col=cell.column, max_col=cell.column):
    #                  for cell_to_link in row_cell:
    #                      cell_to_link.hyperlink = cell_to_link.value
    #              break
    # Sauvegardez le classeur Excel
    #wb.save(chemin)

#---------------------Fonctions pour l'onglets statistiques-------------------------------------

#A partir de "{"feuille":....,"colonne":....}"
#1) extraire la colonne de la feuille
#2)extraire toutes les possibilités de réponses
#3)rentrer dans le dic sous la forme "nom_colonne":[liste_réponses]
def extraction_colonne_ref_possibilité(chemin,feuille,col):
    df = pd.read_excel(chemin, sheet_name=feuille, engine="openpyxl")
    liste_réponses_ref=df[col].unique()
    return {"feuille":feuille,"nom_colonne":col,"liste_possibilité_ref":liste_réponses_ref}

def création_entêtes(listes_réponses_ref):
    liste_entête_NB=[]
    liste_entête_p=[]
    for possibilité in listes_réponses_ref:
        entête_Nb=f"Nb {possibilité}"
        entête_p=f"% {possibilité}"
        liste_entête_NB.append(entête_Nb)
        liste_entête_p.append(entête_p)
    return [liste_entête_NB,liste_entête_p]



def _extraction_réponses_différentes_coll(data,chemin):
    dico=json.loads(data)
    feuille=dico["Feuille"]
    col=dico["Colonne"]
    df=pd.read_excel(chemin,sheet_name=feuille,engine="openpyxl")
    liste_réponses=df[col].unique()
    return [col,liste_réponses]

#Permet extraire les reponses des différentes colonnes dont je veux faire les statistiques
def create_dico_différentes_coll(liste_Qlabel,chemin):
    dico_datas={}
    for Qlabel in liste_Qlabel:
        data=Qlabel.text()
        d=_extraction_réponses_différentes_coll(data,chemin)
        dico_datas[d[0]]=d[1]
    return dico_datas
    #dico_datas={"nom_colonne":[1,2,3],"nom_colonne2":[4,5,6]}




#Toutes les combinaisons des matrices de la forme
# {1: ['a', 'b', 'c'], 2: ['d', 'e', 'f'], 3: ['g', 'h', 'i']}--->
def All_combinaisons_matrice_dictionnary(dictionnaire):

    # Utilisation de product pour générer les combinaisons
    combinaisons = list(product(*dictionnaire.values()))
    liste_combinaison=[]
    #Afficher les combinaisons
    for combinaison in combinaisons:
        c=dict(zip(dictionnaire.keys(), combinaison))
        liste_combinaison.append(c)
    return liste_combinaison
    #liste_combinaison=[["nom_colonne1":1,"nc2":4,"nc3":7],[...]

#fonction pour créer le nom de mes possibilités
def _creation_name_possibility_secondaires(dico_possibility_sec):
    # Initialiser une liste pour stocker les paires nom:valeur
    paires = []

    # Parcourir le dictionnaire et construire les paires
    for nom, valeur in dico_possibility_sec.items():
        paire = f"{valeur}"
        paires.append(paire)

    # Convertir la liste de paires en une seule chaîne de caractères en utilisant '--' comme séparateur
    resultat = '--'.join(paires)
    return resultat

#forme du dictionnaire {col1:{nb1:1,nb2,2...}col2:{nb1:4,nb2:5....}...}--> col1=index,nb1=colonne
def _creation_df_with_dict(dictionnaire):
    # Créer une liste de dictionnaires
    liste_de_dictionnaires = []
    for cle, valeurs in dictionnaire.items():
        ligne = valeurs.copy()
        ligne["index"] = cle
        liste_de_dictionnaires.append(ligne)

    # Créer un DataFrame à partir de la liste de dictionnaires
    df = pd.DataFrame(liste_de_dictionnaires).set_index("index")

    # Remplacer les valeurs None par des NaN si nécessaire
    df = df.fillna(pd.NA)

    # Afficher le DataFrame
    return df





#Creation dico de la forme -> {name_possibilité:{"Nb possibilité1_coll_ref":.....},name_possibilité2:{....}
def compter_possibilité_coll_ref(chemin,feuille,listes_combinaisons_sec,liste_combinaisons_ref):
    df=pd.read_excel(chemin,sheet_name=feuille,engine="openpyxl")
    dico_stats_possibilité={}
    for combi_sec in listes_combinaisons_sec:
        dico_data_combi={}
        name_combi_sec=_creation_name_possibility_secondaires(combi_sec)
        #Variables pour faire les stats % rapidement
        liste_possi_nb_temp=[]
        total=0
        for possibilité_ref in liste_combinaisons_ref["liste_possibilité_ref"]:
            df_filtered = df.copy()
            for colonne, condition in combi_sec.items():
                df_filtered = df_filtered[df_filtered[colonne] == condition]
            df_filtered=df_filtered[df_filtered[liste_combinaisons_ref["nom_colonne"]]==possibilité_ref]
            nb_lignes=len(df_filtered)
            dico_data_combi[f"Nb {possibilité_ref}"]=nb_lignes
            #calcul pourcentage et ajout
            liste_possi_nb_temp.append([possibilité_ref,nb_lignes])
            total+=nb_lignes
        for p in liste_possi_nb_temp:
            if p[1]!=0:
                pourcentage=round((p[1]/total)*100,2)
            else:
                pourcentage=0
            dico_data_combi[f"% {p[0]}"]=pourcentage
        dico_stats_possibilité[name_combi_sec]=dico_data_combi
    df_final=_creation_df_with_dict(dico_stats_possibilité)
    return df_final

#Retourne la liste des widget present dans le scrolle area +liste des différentes possibilités
def extraire_label_dico_scroll_area(scroll_area):
    liste_dico=[]
    liste_widget=[]

    for i in range(scroll_area.count()):
        widget = scroll_area.itemAt(i).widget()
        if widget:
            # Vérifier le type du widget (par exemple, QLineEdit, QLabel, QTextEdit, etc.)
            if isinstance(widget, QLineEdit):
                texte_critère = widget.text()
            elif isinstance(widget, QLabel):
                texte_critère = widget.text()
            elif isinstance(widget, QTextEdit):
                texte_critère = widget.toPlainText()
            liste_dico.append(json.loads(texte_critère))
            liste_widget.append(widget)

    return {'liste_dico':liste_dico,'liste_widget':liste_widget}

#[{Feuille:"eee";"Colonne":"eeee"};{"Feuille":"erre";"Colonne":"..."}]--->{Feuille:[Colonne,Colonne];Feuille:[colo...]}
def _Trie_colonne_per_feuille(dico_feuille):
    # Créez un dictionnaire vide pour stocker les colonnes par feuille
    colonne_par_feuille = {}

# Parcourez la liste de dictionnaires
    for entry in dico_feuille:
        feuille = entry['Feuille']
        colonne = entry['Colonne']

        # Vérifiez si la feuille existe déjà dans le dictionnaire
        if feuille in colonne_par_feuille:
            colonne_par_feuille[feuille].append(colonne)
        else:
            colonne_par_feuille[feuille] = [colonne]
    return colonne_par_feuille

#{Feuille:[Coll,Coll];Feuille:[Coll;Coll]}-------> [df,df,df] avec la colonne Trade intégré
def _return_all_df_Feuille_coll(chemin,liste_dico,Booleen_exclure_datas=False):
    # Créez un dictionnaire pour stocker les DataFrames extraits
    liste_df_colonne=[]
    with open("C:\\Users\\Baptiste\\Documents\\GUI_ext_tri_datas\\données.json","r") as json_file:
        data=json.load(json_file)
    words_trade=data["valeurs_defaults"]["col_ref_words"]
    # Parcourez le dictionnaire colonne_par_feuille
    for feuille, colonnes in liste_dico.items():
        # Ouvrez le fichier Excel et chargez la feuille spécifiée
        df = pd.read_excel(chemin, sheet_name=feuille)
        #Pour n'avoir que la colonne trade existante
        colonne_trade_existante = [colonne for colonne in words_trade if colonne in df.columns]
        colonnes.append(colonne_trade_existante[0])
        # Sélectionnez uniquement les colonnes spécifiées
        df = df[colonnes]
        #Renommage de la colonne Trade
        df.rename(columns={colonne_trade_existante[0]: "Trade"}, inplace=True)

        # Stockez le DataFrame résultant dans le dictionnaire dataframes_par_feuille
        liste_df_colonne.append(df)
    return liste_df_colonne

def _clean_datas(df):
    with open("C:\\Users\\Baptiste\\Documents\\GUI_ext_tri_datas\\données.json","r") as json_file:
        data=json.load(json_file)
    mots_exclure=data["valeurs_defaults"]["words_exclure_df"]
    for colonne in df.columns:
        df = df[~(df[colonne].isin(mots_exclure) | (df[colonne] == '') | df[colonne].isna())]
    return df

def _fusion_liste_df(liste_df):
    merged_df = pd.concat(liste_df, axis=0, ignore_index=True)
    return merged_df

def _create_df_state_no_coll_ref(df):
    pass

#reponse-> {nom col:[rpe,rep,rep];nom col:[rep,rep,rep]...}
def extration_all_possibilities_without_coll_trade(df):
    reponses={}
    # Parcourez les colonnes du DataFrame
    for colonne in df.columns:
        # Obtenez les valeurs uniques de la colonne (en excluant les valeurs vides)
        valeurs_uniques = df[colonne][df[colonne] != ''].unique()

        # Ajoutez les valeurs uniques à votre dictionnaire sous la clé de la colonne
        reponses[colonne] = list(valeurs_uniques)
    return reponses

def fusion_clean_dicos(chemin,liste_dico_feuille,Boolean_exlure_datas):
    dico_Feuille_coll=_Trie_colonne_per_feuille(liste_dico_feuille)
    liste_df=_return_all_df_Feuille_coll(chemin,dico_Feuille_coll)
    df=_fusion_liste_df(liste_df)
    if Boolean_exlure_datas:
        df=_clean_datas(df)
    return df

def comptage_possibilities_df(df,liste_dicos_possibilities):
# Créez un dictionnaire pour stocker le nombre de lignes par possibilité
    comptage_possibilites = {}
    # Parcourez les possibilités et comptez le nombre de lignes correspondant à chaque possibilité
    for possibilite in liste_dicos_possibilities:
        query_string = ' and '.join([f'`{col}` == "{val}"' for col, val in possibilite.items()])
        matching_rows = df.query(query_string)
        counter = len(matching_rows)
        comptage_possibilites['-'.join(possibilite.values())] = counter
    return comptage_possibilites

#{name:Nombre;name:Nombre}
def creation_df_stat_ss_col_ref(dicos_possibility):
    pass
def somme_values_dico(dictionnaire):
    # Parcourez les valeurs du dictionnaire et ajoutez-les à la somme
    somme=0
    for valeur in dictionnaire.values():
        try:
            somme += int(valeur)
        except ValueError:pass
    return somme
def _creation_df_stat_ss_col_ref(possibility_nb,nb_lignes):
    dico_final={}
    for key,value in possibility_nb.items():
        pourcentage=round((value/nb_lignes)*100,2)
        dico_final[key]={'Nombre':value,'Pourcentage%':pourcentage}
    df=pd.DataFrame(dico_final).T
    return df

def stats_without_col_ref(liste_dicos_possibilities,df):
    possibility_nb=comptage_possibilities_df(df,liste_dicos_possibilities)
    nb_lignes=somme_values_dico(possibility_nb)
    df=_creation_df_stat_ss_col_ref(possibility_nb,nb_lignes)
    return df


    #Calculer les stats

    #Je dois créer les dataframmes des différentes feuilles
    #liste dico contient la feuille et les colonnes :
# [{'Feuille': '4H', 'Colonne': 'tendance/contretendance 4H'}, {'Feuille': '4H', 'Colonne': 'Type Poi'}]
    #=>Classer les dicos en fonctions de leur feuilles, creer un dataframme avec les colonnes en laissant la colonne Trade dans les feuilles
    #merger les différents df avec la fonction merge colonne Trade effacer les lignes avec les colonnes vides si la case est coché
    #


def création_sheet_excel(chemin,df,name_feuille=""):
    if name_feuille=="":
        workbook = openpyxl.load_workbook(chemin)
        name_feuille="Statistiques_feuille"


    # Ouvrir un fichier Excel existant
    with pd.ExcelWriter(chemin, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
        # Écrire le DataFrame dans une nouvelle feuille
        df.to_excel(writer, sheet_name=name_feuille)



#------------------------------Fonctions onglet 3 traitements
def scroll_area_json_to_list_dico(widget_scroll):
    liste_critères = []
    for i in range(widget_scroll.count()):
        widget = widget_scroll.itemAt(i).widget()
        if widget:
            # Vérifier le type du widget (par exemple, QLineEdit, QLabel, QTextEdit, etc.)
            if isinstance(widget, QLineEdit):
                texte_critère = widget.text()
            elif isinstance(widget, QLabel):
                texte_critère = widget.text()
            elif isinstance(widget, QTextEdit):
                texte_critère = widget.toPlainText()
            liste_critères.append(json.loads(texte_critère))
    return liste_critères

def find_colonne_trade(df):
    with open("C:\\Users\\Baptiste\\Documents\\GUI_ext_tri_datas\\données.json", "r") as json_file:
        datas = json.load(json_file)
    word_trade = datas["valeurs_defaults"]["col_ref_words"]
    for word in word_trade:
        try:
            df_trade=df[[word]].copy()
            df_trade.columns=["Trade"]
            return df_trade
        except:pass
    return False
def _extraction_colonne(chemin,dico_colonne,list_value_trade_colonne):

    feuille=dico_colonne["Feuille"]
    colonne=dico_colonne["Colonne"]
    df = pd.read_excel(chemin, sheet_name=feuille)
    for word in list_value_trade_colonne:
        try:
            df_trade=df[word].copy()
            df_trade.columns=["Trade"]
        except:pass
    df_main_colonne=df[colonne]
    df=pd.concat([df_trade,df_main_colonne],axis=1)
    return df



def extration_colonnes(chemin,liste_dico_colonne):
    with open("C:\\Users\\Baptiste\\Documents\\GUI_ext_tri_datas\\données.json","r") as json_file:
        datas=json.load(json_file)
    word_trade=datas["valeurs_defaults"]["col_ref_words"]
    liste_colonne_df=[]
    for dico in liste_dico_colonne:
        df_coll=_extraction_colonne(chemin,dico,word_trade)
        liste_colonne_df.append(df_coll)
    df = pd.concat(liste_colonne_df, ignore_index=True)
    df = df.groupby("Trade").sum().reset_index()
    return df

def ecriture_df_to_excel_col_suivante(chemin,nom_feuille,df_a_ajouter):
    book=openpyxl.load_workbook(chemin)
    feuille=book[nom_feuille]
    derniere_colonne = feuille.max_column if feuille.max_column is not None else 0
    for idx, col in enumerate(df_a_ajouter.columns):
        feuille.cell(row=1, column=derniere_colonne + idx + 1, value=col)
        for i, val in enumerate(df_a_ajouter[col]):
            feuille.cell(row=i + 2, column=derniere_colonne + idx + 1, value=val)

        # Enregistrez le fichier Excel
    book.save(chemin)


#1 regarder quel est la derniere colonne
#1.2=> ouvrir un df pour savoir à partir de quel col est la fin en nombre de colonne
#2inserer à la suite le df avec openpyxl les values, en les inserer une par une













